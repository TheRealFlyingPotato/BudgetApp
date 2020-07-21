import openpyxl
import sys
import datetime
import backup
import os
import funcs

FIN = funcs.FIN
FOUT = funcs.FOUT
COLS = funcs.COLS
DATE_COL = funcs.DATE_COL
STARTING_ROW = funcs.STARTING_ROW
ITEM_COL = ITEM_COL=funcs.ITEM_COL

DATE = datetime.datetime.now().strftime("%m/%d/%Y")

def parse_sheet(sheet):
  return funcs.parse_sheet(sheet)

def find_open_row(sheet):
  return funcs.find_open_row(sheet)

def generate_monetary_values(sheet_data, amt):
  column_vals = dict()
  for key, val in sheet_data['column_attributes'].items():
    # print(key, val)
    if val['max'] == None:
      column_vals[key] = val['default']
    else:
      column_vals[key] = min(val['max'] - val['current'], val['default'])
    if column_vals[key] > amt:
      column_vals[key] = amt
    if column_vals[key] < 0:
      column_vals[key] = 0
    amt -= column_vals[key]
    if amt <= 0:
      break
  # increment_triggered = False
  while amt > 0:
    for key, val in sheet_data['column_attributes'].items():
      # print(f'{key}, {val} > remaining: {amt}')
      if val['increment'] != None:
        addition = min(amt, val['increment'])
        if val['max'] != None:
          addition = min(addition, val['max'] - (val['current'] + column_vals[key]))
          if addition < 0:
            addition = 0
        column_vals[key] += addition
        amt -= addition
      if amt == 0:
        break
  column_vals[sheet_data['cushion_col']] += amt

  return column_vals

def set_sheet_vals(sheet, vals, title, row):
  funcs.set_sheet_vals(sheet, vals, title, row)

def cleanup(filename):
  funcs.cleanup(filename)

def main(title, amt, outputHandle=None):
  backup.backup(FOUT)
  wb0 = openpyxl.load_workbook(filename = FIN, data_only=True)
  sheet_data = parse_sheet(wb0[wb0.sheetnames[0]])
  wb0.close()


  wb = openpyxl.load_workbook(filename=FIN)
  sheet = wb[wb.sheetnames[0]]
  open_row = find_open_row(sheet)
  column_vals = generate_monetary_values(sheet_data, amt)
  set_sheet_vals(sheet, column_vals, title, open_row)

  wb.save(FOUT)
  print(f"Deposited {amt} as {title}")
  if outputHandle != None:
    outputHandle.output(f"Deposited {amt} as {title}", line=False)

  cleanup(FIN)


if __name__ == "__main__":
  try:
    amt = int(sys.argv[1])
  except:
    raise Exception("No amount given")
  try:
    title = sys.argv[2]
  except:
    title = 'Paycheck'
  main(title, amt)

# wb = openpyxl.load_workbook(filename = 'test.xlsx', data_only=True)
# print("Finding Open Row")
# sheet = wb[wb.sheetnames[0]]
# open_row = find_open_row(sheet)
# print(f'Row Found: {sheet["B" + str(open_row)].value} >> {open_row}')
# wb.close()
# wb2 = openpyxl.load_workbook(filename = 'test.xlsx')
# sheet2 = wb2[wb2.sheetnames[0]]
# # print(sheet2['B8'].value)
# sheet2['C' + str(open_row)] = 42069
# # print([sheet2['C3'].value])
# print([sheet2['B8'].value])
# wb2.save("test.xlsx")