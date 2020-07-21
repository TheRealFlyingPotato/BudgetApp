import funcs
import backup
import openpyxl
from prettytable import PrettyTable

def main(outputHandler = None):
  backup.backup(funcs.FOUT)
  workbook = openpyxl.load_workbook(filename = funcs.FOUT, data_only=True)
  sheet = funcs.fetch_sheet(workbook)
  sheet_data = funcs.parse_sheet(sheet)
  workbook.close()
  st = ''
  tbl = PrettyTable()
  tbl.field_names = ['Item', 'USD']
  st += f"Total\t:{sheet_data['total_cash']}\n"
  tbl.add_row(['$', round(sheet_data['total_cash'], 2)])
  for _, val in sheet_data['column_attributes'].items():
    st += f"{val['name']}\t:{val['current']}\n"
    tbl.add_row([val['name'], round(val['current'], 2)])
  st = st.strip()
  if outputHandler != None:
    # outputHandler.output(st)
    outputHandler.output(str(tbl))
  else:
    # print(st)
    print(tbl)

if __name__ == "__main__":
  main()

# import funcs
# import backup
# import openpyxl
# from prettytable import PrettyTable

# def main(outputHandler = None):
#   backup.backup(funcs.FOUT)
#   wb0 = openpyxl.load_workbook(filename = funcs.FIN, data_only=True)
#   sheet_data = funcs.parse_sheet(wb0[wb0.sheetnames[0]])
#   wb0.close()
#   st = ''
#   tbl = PrettyTable()
#   tbl.field_names = ['Item', 'USD']
#   for _, val in sheet_data['column_attributes'].items():
#     st += f"{val['name']}\t:{val['current']}\n"
#     tbl.add_row([val['name'], round(val['current'], 2)])
#   st = st.strip()
#   if outputHandler != None:
#     # outputHandler.output(st)
#     outputHandler.output(str(tbl))
#   else:
#     # print(st)
#     print(tbl)

# if __name__ == "__main__":
#   main()